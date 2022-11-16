
#* importamos la libreria para poder conectanos 
from openpyxl import Workbook, load_workbook
import requests
from json import dumps

sheet = r'C:\Users\jbsot\Documents\Curso Python\Ejercicios\jholder.xlsx'

web = requests.get('https://jsonplaceholder.typicode.com/users')
data = web.json() #* le doy formato json
n_data = dumps(data, indent= 4)


def crear_libro(): #! create
    libro = Workbook() #! creamos el archivo de excel 
    hoja = libro.active #! activamos una hoja del archivo 
    hoja.title = 'JsonHolder' #! damos nombre a la hoja 
    encabezado = ['id', 'name', 'username', 'email', 'address', 'phone', 'website', 'company'] #! pasamos una lista para el encabezado 
    hoja.append(encabezado) #! para cargar el encabezado debemos hacer un append 
    libro.save(filename= sheet) #! le indicamos la ruta
    #! el nombre y la extension del archivo, por ultimo guardamos 
    libro.close() #! cerramos elñ archivo 


def agregar_datos(): #! update
    libro = load_workbook(filename= sheet) #* abrimos el libro 
    hoja = libro['JsonHolder']  #* se puede indicar de igual forma como libro.active
    for datos in data: #* realizamos un for para recorrer las lista que traemos con el get 
        hoja.append([datos['id'], datos['name'], datos['email'], datos['phone'] , datos['website']]) #* con .append cargados los datos que necesitamos 
        #* indicando el key
    libro.save(filename= sheet)
    libro.close()











#### de manera local #####


#print(data[0]['id'])


#for keys in data:
#   for valor in keys:



